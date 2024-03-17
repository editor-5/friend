from selenium import webdriver

driver = webdriver.Chrome(r'C:\Users\gas00\MathTool\codeit\web_automation\chromedriver-win64')
driver.implicitly_wait(3)

driver.get('https://workey.codeit.kr/costagram')
login_link = driver.find_element_by_css_selector('.top-nav__login-link')
login_link.click()


'''
# Selenium 임포트
from selenium import webdriver

# 크롬 드라이버 생성
driver = webdriver.Chrome(r"C:\Users\gas00\MathTool\codeit\web_automation\chromedriver-win64")

# 사이트 접속하기
driver.get('https://codeit.kr')


import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook(write_only=True)

for year in range(2010, 2019):
    ws = wb.create_sheet("{}년".format(year))
    ws.append(['기간', '순위', '채널', '프로그램', '시청율'])
    for month in range(1, 13):
        for weekIndex in range(0, 5):
            url = "https://workey.codeit.kr/ratings/index?year={}&month={}&weekIndex={}".format(year, month, weekIndex)
            response = requests.get(url)
            rating_page = response.text
            soup = BeautifulSoup(rating_page, 'html.parser')

            for tr_tag in soup.select('tr')[1:]:
                td_tags = tr_tag.select('td')
                period = "{}년 {}월 {}주차".format(year, month, weekIndex + 1)
                row = [
                    period,
                    td_tags[0].get_text(),
                    td_tags[1].get_text(),
                    td_tags[2].get_text(),
                    td_tags[3].get_text(),
                ]
                ws.append(row)
wb.save('시청률.xlsx')


import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook(write_only=True)
ws = wb.create_sheet(('TV Ratings'))
ws.append(['순위', '채널', '프로그램', '시청율'])

response = requests.get("https://workey.codeit.kr/ratings/index")
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')

for tr_tag in soup.select('tr')[1:]:
    td_tags = tr_tag.select('td')
    row = [
        td_tags[0].get_text(),
        td_tags[1].get_text(),
        td_tags[2].get_text(),
        td_tags[3].get_text(),
    ]
    ws.append(row)
wb.save('시청률_2010년1월1주차.xlsx')

import requests
from bs4 import BeautifulSoup

response = requests.get("https://workey.codeit.kr/music")
music_page = response.text

soup = BeautifulSoup(music_page, 'html.parser')
popular_artists = []
for tag in soup.select('ul.popular__order li'):
    popular_artists.append(list(tag.stripped_strings)[1])

print(popular_artists)


import requests
from bs4 import BeautifulSoup

response = requests.get("https://workey.codeit.kr/music")
music_page = response.text

soup = BeautifulSoup(music_page, 'html.parser')

# CSS 선택자 수정
popular_artists = []
for tag in soup.select('ul.popular__order li'):
    popular_artists.append(tag.get_text().strip())
print(popular_artists)


import requests
from bs4 import BeautifulSoup

response = requests.get("https://workey.codeit.kr/ratings/index")
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')

tr_tag = soup.select('tr')[1]
td_tags = tr_tag.select('td')
print(td_tags)

for tag in td_tags:
    print(tag.get_text())

import requests
from bs4 import BeautifulSoup

response = requests.get("https://workey.codeit.kr/ratings/index")
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')


td_tags = soup.select('td')[:4]

for tag in td_tags:
    print(tag.get_text())


import requests
from bs4 import BeautifulSoup

response = requests.get("https://workey.codeit.kr/ratings/index")
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')

program_title_tags = soup.select('td.program')

print(soup.select_one('td.program'))




import requests
from bs4 import BeautifulSoup

response = requests.get("https://workey.codeit.kr/ratings/index")
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')

program_title_tags = soup.select('td.program')

program_titles = []
for tag in program_title_tags:
    program_titles.append(tag.get_text())

print(program_titles)    

'''




